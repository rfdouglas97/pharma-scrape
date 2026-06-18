"""Parse downloadable pipeline documents (CSV / XLSX / PDF) into clean text.

Many companies publish their pipeline as a downloadable file that is cleaner and more
complete than the rendered page (GSK's quarterly spreadsheet, table-based PDFs). We parse
those into Markdown-rendered tables + text so the structure survives into the existing
text-only LLM extraction path, and the text is used for content-hash change detection.

This is artifact-specific, not company-specific: a CSV is parsed as a CSV regardless of
which company published it.
"""

from __future__ import annotations

import csv as _csv
import io
from dataclasses import dataclass, field

# Map a content-type or file extension to a parser kind.
_CSV = {"text/csv", "application/csv", ".csv", "csv_doc"}
_XLSX = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    ".xlsx",
    ".xls",
    "xlsx_doc",
}
_PDF = {"application/pdf", ".pdf", "pdf_doc"}

# A parsed table with this many rows or more keeps its full Markdown rendering; tiny
# tables are still rendered (they may be the whole pipeline on a small-cap company).
MAX_PDF_PAGES = 50


@dataclass
class ParsedDoc:
    """`text` is the authoritative extraction/hashing input; `tables` is the structured
    fallback (one entry per source table, rows of string cells)."""

    text: str
    kind: str  # csv | xlsx | pdf
    tables: list[list[list[str]]] = field(default_factory=list)
    n_pages: int | None = None


def doc_kind(content_type: str | None = None, ext: str | None = None) -> str | None:
    """Resolve a parser kind from a content-type and/or extension. Returns None if neither
    identifies a supported document type."""
    for token in (content_type, ext):
        if not token:
            continue
        t = token.strip().lower().split(";", 1)[0]
        if t in _CSV:
            return "csv"
        if t in _XLSX:
            return "xlsx"
        if t in _PDF:
            return "pdf"
    return None


def parse_document(raw: bytes, content_type: str | None = None, ext: str | None = None) -> ParsedDoc:
    kind = doc_kind(content_type, ext)
    if kind == "csv":
        return _parse_csv(raw)
    if kind == "xlsx":
        return _parse_xlsx(raw)
    if kind == "pdf":
        return _parse_pdf(raw)
    raise ValueError(f"unsupported document type (content_type={content_type!r}, ext={ext!r})")


def render_markdown_table(rows: list[list[str]]) -> str:
    """Render rows of cells as a GitHub-flavored Markdown table. The first row is treated
    as the header. Empty input -> empty string."""
    rows = [["" if c is None else str(c).strip() for c in row] for row in rows if any(row)]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header, *body = rows
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    lines += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(lines)


def _parse_csv(raw: bytes) -> ParsedDoc:
    decoded = raw.decode("utf-8-sig", errors="replace")
    try:
        dialect = _csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t|")
    except _csv.Error:
        dialect = _csv.excel
    rows = [list(r) for r in _csv.reader(io.StringIO(decoded), dialect)]
    return ParsedDoc(text=render_markdown_table(rows), kind="csv", tables=[rows] if rows else [])


def _parse_xlsx(raw: bytes) -> ParsedDoc:
    from openpyxl import load_workbook  # noqa: PLC0415 — heavy import, defer

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    tables: list[list[list[str]]] = []
    parts: list[str] = []
    for ws in wb.worksheets:
        rows = [
            ["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(r)]
        if not rows:
            continue
        tables.append(rows)
        parts.append(f"## Sheet: {ws.title}\n\n{render_markdown_table(rows)}")
    wb.close()
    return ParsedDoc(text="\n\n".join(parts), kind="xlsx", tables=tables)


def _parse_pdf(raw: bytes) -> ParsedDoc:
    import pdfplumber  # noqa: PLC0415 — heavy import, defer

    tables: list[list[list[str]]] = []
    parts: list[str] = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        n_pages = len(pdf.pages)
        for idx, page in enumerate(pdf.pages[:MAX_PDF_PAGES], start=1):
            page_parts: list[str] = []
            for tbl in page.extract_tables() or []:
                rows = [["" if c is None else str(c).strip() for c in row] for row in tbl]
                if any(any(r) for r in rows):
                    tables.append(rows)
                    page_parts.append(render_markdown_table(rows))
            text = (page.extract_text() or "").strip()
            if text:
                page_parts.append(text)
            if page_parts:
                parts.append(f"### Page {idx}\n\n" + "\n\n".join(page_parts))
    return ParsedDoc(text="\n\n".join(parts), kind="pdf", tables=tables, n_pages=n_pages)
